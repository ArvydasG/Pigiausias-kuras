import os
import sys
import json
import time
import subprocess
import random
import datetime

def ensure_packages():
    required = {"requests", "beautifulsoup4", "openpyxl"}
    try:
        import pkg_resources
        installed = {pkg.key for pkg in pkg_resources.working_set}
        missing = required - installed
        if missing:
            print(f"Diegiami trukstami paketai: {missing}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])
    except Exception as e:
        print(f"Nepavyko patikrinti paketu: {e}. Bandoma tesi...")

ensure_packages()

import requests
from bs4 import BeautifulSoup
import openpyxl

DATA_FILE = "data.js"
CACHE_FILE = "coords_cache.json"

# Load coordinate cache
coords_cache = {}
if os.path.exists(CACHE_FILE):
    with open(CACHE_FILE, "r", encoding="utf-8") as f:
        coords_cache = json.load(f)

def save_cache():
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(coords_cache, f, ensure_ascii=False, indent=4)

def geocode_address(address, city):
    """Fetches coordinates for a given address using ArcGIS Geocoding API."""
    clean_address = address.replace('\xa0', ' ').replace('\u200b', '').replace(chr(0x200B), '').strip()
    search_query = f"{clean_address}, {city}, Lietuva"
    if search_query in coords_cache:
        return coords_cache[search_query]
        
    print(f"Ieskoma koordinaciu su ArcGIS: {search_query.encode('cp1257', 'ignore').decode('cp1257')}...")
    
    url = "https://geocode.arcgis.com/arcgis/rest/services/World/GeocodeServer/findAddressCandidates"
    params = {
        "SingleLine": search_query,
        "f": "json",
        "maxLocations": 1,
        "outFields": "Match_addr"
    }
    
    try:
        time.sleep(0.4) # Trumpesnis laukimas, nes ArcGIS greitas
        res = requests.get(url, params=params).json()
        
        if res.get('candidates') and len(res['candidates']) > 0:
            location = res['candidates'][0]['location']
            coords = {"lat": float(location['y']), "lng": float(location['x'])}
            coords_cache[search_query] = coords
            return coords
        else:
            print(f"NERASTA: {search_query}. Metama į miesto centrą su triukšmu...")
            time.sleep(0.4)
            res_city = requests.get(url, params={"SingleLine": f"{city}, Lietuva", "f": "json", "maxLocations": 1}).json()
            if res_city.get('candidates'):
                loc = res_city['candidates'][0]['location']
                lat_noise = random.uniform(-0.005, 0.005)
                lng_noise = random.uniform(-0.005, 0.005)
                coords = {"lat": float(loc['y']) + lat_noise, "lng": float(loc['x']) + lng_noise}
                coords_cache[search_query] = coords
                return coords
            
    except Exception as e:
        print(f"Klaida ieškant koordinačių su ArcGIS: {e}")
    
    return {"lat": 54.6872, "lng": 25.2797}

def fetch_data():
    default_discounts = {
        "Circle K": 0.035,
        "Neste": 0.035,
        "Viada": 0.030,
        "Baltic Petroleum": 0.000,
        "Emsi": 0.000,
        "Jozita": 0.000,
        "Saurida": 0.000,
        "Orlen": 0.000
    }
    print("Pradedamas degalų kainų duomenų siuntimas iš LEA (ena.lt)...")
    
    # 1. Fetch ENA homepage to find SharePoint link
    res = requests.get("https://www.ena.lt/degalu-kainos-degalinese/")
    soup = BeautifulSoup(res.text, "html.parser")
    
    links = soup.find_all("a", href=True)
    excel_url = None
    for a in links:
        if "sharepoint.com" in a["href"] and ("Naujausios" in a.text or "Degalų kainos" in a.get("title", "")):
            excel_url = a["href"]
            break
            
    if not excel_url:
        print("Klaida: Nerasta nuoroda į LEA Excel failą.")
        return
        
    print(f"Rasta Excel nuoroda: {excel_url}")
    
    # 2. Add download flag for Sharepoint
    if "?" in excel_url:
        download_url = excel_url + "&download=1"
    else:
        download_url = excel_url + "?download=1"
        
    # 3. Download the Excel file
    print("Siunčiamas Excel failas...")
    excel_data = requests.get(download_url).content
    temp_file = "temp_lea.xlsx"
    with open(temp_file, "wb") as f:
        f.write(excel_data)
        
    # 4. Parse Excel file using openpyxl
    print("Skaitomas Excel failas...")
    wb = openpyxl.load_workbook(temp_file, data_only=True)
    sheet = wb.active
    
    stations_dict = {}
    
    # Find the header row index
    header_row_idx = None
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row[0] and str(row[0]).strip().lower() == "įmonė":
            header_row_idx = idx
            break
            
    if not header_row_idx:
        print("Nepavyko rasti duomenų antraštės Excel faile.")
        return
        
    # Iterate through rows below the header
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row[0]: # No company name means end of data or empty row
            continue
            
        name = str(row[0]).strip()
        raw_city = str(row[1]).strip()
        city = raw_city.replace(" m. sav.", "").replace(" r. sav.", "").replace(" sav.", "")
        address = str(row[2]).strip()
        fuel_type_raw = str(row[3]).strip()
        
        def extract_price(val):
            if val is None:
                return None
            try:
                if isinstance(val, str):
                    val = val.replace(',', '.').strip()
                    if val.lower() == 'neprekiauja' or val == '-' or not val:
                        return None
                return float(val)
            except ValueError:
                return None
                
        price = extract_price(row[4])
        if price is None:
            continue
            
        key = (name, raw_city, address)
        if key not in stations_dict:
            display_city = city.replace("Vilniaus", "Vilnius").replace("Kauno", "Kaunas").replace("Klaipėdos", "Klaipėda")
            
            # Geocode using the raw city name (e.g. 'Kauno r. sav.') so ArcGIS doesn't confuse it with the city center
            coords = geocode_address(address, raw_city)
                
            # Determine logo emoji
            logo = "⛽"
            name_lower = name.lower()
            if "circle k" in name_lower: logo = "🔴"
            elif "viada" in name_lower: logo = "🦌"
            elif "neste" in name_lower: logo = "🟢"
            elif "baltic petroleum" in name_lower: logo = "🔵"
            elif "emsi" in name_lower: logo = "🛢️"
            elif "jozita" in name_lower: logo = "🟡"
            elif "saurida" in name_lower: logo = "🔥"
            elif "orlen" in name_lower: logo = "🦅"
                
            stations_dict[key] = {
                "name": name,
                "logo": logo,
                "city": display_city,
                "address": address,
                "lat": coords["lat"],
                "lng": coords["lng"],
                "prices": {
                    "A95": None,
                    "A98": None,
                    "Diesel": None,
                    "LPG": None
                }
            }
            
        if fuel_type_raw == '95 benzinas':
            stations_dict[key]["prices"]["A95"] = price
        elif fuel_type_raw == 'Dyzelinas':
            stations_dict[key]["prices"]["Diesel"] = price
        elif fuel_type_raw == 'SND':
            stations_dict[key]["prices"]["LPG"] = price

    stations = []
    count = 0
    for key, st in stations_dict.items():
        st["id"] = count + 1
        stations.append(st)
        count += 1
        # Save cache every 10 items
        if count % 10 == 0:
            save_cache()
            
    # Cleanup temp file
    if os.path.exists(temp_file):
        os.remove(temp_file)
        
    save_cache()
    
    # Final Save to data.js
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    js_content = f"// Automatiškai sugeneruoti duomenys iš LEA Excel\nconst lastUpdated = '{now_str}';\nconst defaultDiscounts = {json.dumps(default_discounts, indent=4, ensure_ascii=False)};\nconst stationsData = {json.dumps(stations, indent=4, ensure_ascii=False)};"
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        f.write(js_content)
        
    print(f"\nSėkmingai išsaugota {len(stations)} degalinių faile {DATA_FILE}!")

if __name__ == "__main__":
    fetch_data()
